"""点検結果のエクスポート(Excel / CSV)"""

from __future__ import annotations

import csv
import io
import re

from ..engine import CheckResult
from ..models import Severity

SEVERITY_LABEL = {
    Severity.ERROR: "エラー",
    Severity.WARNING: "警告",
    Severity.INFO: "情報",
}

HEADERS = [
    "レセプト番号", "患者氏名", "重大度", "分類", "ルールID", "ルール名",
    "指摘内容", "対象", "根拠・補足", "対処提案",
]

# Excelで許容されない制御文字
_ILLEGAL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _safe_cell(v):
    """セル値の無害化: 制御文字の除去と数式解釈(=,+,-,@始まり)の防止"""
    if not isinstance(v, str):
        return v
    v = _ILLEGAL_CHARS.sub("", v)
    if v[:1] in ("=", "+", "@") or (v[:1] == "-" and len(v) > 1 and not v[1:2].isspace()):
        return "'" + v
    return v


def _rows(result: CheckResult):
    for f in result.findings:
        yield [
            f.receipt_no or "",
            _safe_cell(f.patient_name),
            SEVERITY_LABEL[f.severity],
            f.category,
            f.rule_id,
            f.rule_name,
            _safe_cell(f.message),
            _safe_cell(f.target),
            _safe_cell(f.detail),
            _safe_cell(f.suggestion),
        ]


def to_csv_bytes(result: CheckResult) -> bytes:
    """CSV(Excel互換のためBOM付きUTF-8)"""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(HEADERS)
    for row in _rows(result):
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")


def to_excel_bytes(result: CheckResult) -> bytes:
    """Excelブック(サマリー+指摘一覧)"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # --- サマリーシート ---
    ws = wb.active
    ws.title = "サマリー"
    ws["A1"] = "レセプト点検結果サマリー"
    ws["A1"].font = Font(size=14, bold=True)
    f = result.claim_file.facility
    summary = [
        ("医療機関", f.name or f.facility_code),
        ("請求年月", f.seikyu_ym),
        ("ファイル", result.claim_file.source_name),
        ("レセプト件数", len(result.claim_file.receipts)),
        ("指摘のあるレセプト", result.receipts_with_findings),
        ("エラー", result.error_count),
        ("警告", result.warning_count),
        ("情報", result.info_count),
        ("実行ルール数", result.rules_run),
    ]
    for i, (label, value) in enumerate(summary, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=_safe_cell(value))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40

    # 分類別集計
    row0 = len(summary) + 4
    ws.cell(row=row0, column=1, value="分類別件数").font = Font(bold=True)
    for i, (cat, items) in enumerate(sorted(result.by_category().items()), start=row0 + 1):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=len(items))

    # --- 指摘一覧シート ---
    ws2 = wb.create_sheet("指摘一覧")
    fill = {
        "エラー": PatternFill("solid", fgColor="FFC7CE"),
        "警告": PatternFill("solid", fgColor="FFEB9C"),
        "情報": PatternFill("solid", fgColor="DDEBF7"),
    }
    for col, h in enumerate(HEADERS, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
    widths = [10, 14, 8, 12, 9, 26, 60, 30, 40, 40]
    for col, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    for r, row in enumerate(_rows(result), start=2):
        for col, v in enumerate(row, start=1):
            c = ws2.cell(row=r, column=col, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=col >= 6)
        sev = row[2]
        if sev in fill:
            ws2.cell(row=r, column=3).fill = fill[sev]
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(2, ws2.max_row)}"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
